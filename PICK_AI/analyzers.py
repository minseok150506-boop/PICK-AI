import csv
import json
import shutil
import tempfile
import uuid
from pathlib import Path

from config import (
    ALLOWED_IMAGE_EXT, ALLOWED_VIDEO_EXT, ALLOWED_DOC_EXT,
    MAX_UPLOAD_MB, UPLOAD_DIR
)
from pick_llm import vision_analyze, PickLLMRouter

try:
    from PIL import Image
except Exception:
    Image = None

try:
    import cv2
except Exception:
    cv2 = None

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

try:
    from docx import Document
except Exception:
    Document = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def _safe_upload(file_storage, allowed_ext):
    original = Path(file_storage.filename or "upload.bin")
    ext = original.suffix.lower()
    if ext not in allowed_ext:
        raise ValueError(f"지원하지 않는 파일 형식입니다: {ext or '확장자 없음'}")

    target = UPLOAD_DIR / f"{uuid.uuid4().hex}{ext}"
    file_storage.save(target)

    max_bytes = MAX_UPLOAD_MB * 1024 * 1024
    if target.stat().st_size > max_bytes:
        target.unlink(missing_ok=True)
        raise ValueError(f"파일 크기는 최대 {MAX_UPLOAD_MB}MB까지 지원합니다.")
    return target, original.name


def analyze_image_upload(file_storage):
    path, original = _safe_upload(file_storage, ALLOWED_IMAGE_EXT)
    try:
        base = {"filename": original, "size_bytes": path.stat().st_size}
        if Image:
            with Image.open(path) as img:
                base.update({
                    "width": img.width,
                    "height": img.height,
                    "format": img.format,
                    "mode": img.mode,
                })
        try:
            description = vision_analyze(
                [path],
                "이 이미지를 한국어로 자세히 분석해 주세요. 보이는 객체와 장면을 설명하고, 이미지 속 글자는 가능한 한 원문 그대로 정확히 옮겨 적으세요. 판독이 어려운 글자는 추측하지 말고 판독 불가라고 표시해 주세요."
            )
            base["analysis"] = description
            base["vision_model_used"] = True
        except Exception as exc:
            base["analysis"] = (
                "비전 모델 분석을 실행하지 못했습니다. "
                "PICK_VISION_MODEL(기본 llava:latest)이 미니PC Ollama에 설치되어 있는지 확인해 주세요."
            )
            base["vision_model_used"] = False
            base["vision_error"] = str(exc)
        return base
    finally:
        path.unlink(missing_ok=True)


def translate_image_upload(file_storage, target_language="한국어"):
    path, original = _safe_upload(file_storage, ALLOWED_IMAGE_EXT)
    try:
        base = {"filename": original, "size_bytes": path.stat().st_size, "target_language": target_language}
        if Image:
            with Image.open(path) as img:
                base.update({"width": img.width, "height": img.height, "format": img.format, "mode": img.mode})
        prompt = f"""이 이미지에 보이는 글자를 가능한 한 정확히 읽고 {target_language}로 번역해 주세요.
규칙:
1. 먼저 [원문] 아래에 보이는 글자를 줄 순서대로 최대한 정확히 옮기세요.
2. 그 다음 [번역] 아래에 같은 순서로 {target_language} 번역을 적으세요.
3. 읽을 수 없는 부분은 추측하지 말고 [판독 불가]라고 표시하세요.
4. 숫자, 날짜, 단위, 고유명사, URL, 코드, 제품명은 함부로 바꾸지 마세요.
5. 표/메뉴/버튼이면 원래 구조가 이해되도록 항목 순서를 유지하세요.
6. 설명보다 원문과 번역 결과를 우선하세요.
"""
        try:
            base["analysis"] = vision_analyze([path], prompt)
            base["vision_model_used"] = True
            base["translation"] = True
        except Exception as exc:
            base["analysis"] = "이미지 번역을 실행하지 못했습니다. PICK 비전 모델 상태를 확인해 주세요."
            base["vision_model_used"] = False
            base["vision_error"] = str(exc)
        return base
    finally:
        path.unlink(missing_ok=True)


def _extract_video_frames(path, count=6):
    if cv2 is None:
        raise RuntimeError("opencv-python-headless가 설치되어 있지 않습니다.")

    cap = cv2.VideoCapture(str(path))
    try:
        total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
        fps = float(cap.get(cv2.CAP_PROP_FPS) or 0)
        duration = (total / fps) if total and fps else None
        frames = []
        if total <= 0:
            positions = [0]
        else:
            positions = [int(i * max(total - 1, 0) / max(count - 1, 1)) for i in range(count)]

        tmpdir = Path(tempfile.mkdtemp(prefix="pick_video_"))
        for idx, pos in enumerate(positions):
            cap.set(cv2.CAP_PROP_POS_FRAMES, pos)
            ok, frame = cap.read()
            if not ok:
                continue
            frame_path = tmpdir / f"frame_{idx:02d}.jpg"
            cv2.imwrite(str(frame_path), frame)
            frames.append(frame_path)

        return frames, {
            "frame_count": total,
            "fps": round(fps, 2) if fps else None,
            "duration_seconds": round(duration, 2) if duration else None,
            "width": int(cap.get(cv2.CAP_PROP_FRAME_WIDTH) or 0),
            "height": int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT) or 0),
        }, tmpdir
    finally:
        cap.release()


def analyze_video_upload(file_storage):
    path, original = _safe_upload(file_storage, ALLOWED_VIDEO_EXT)
    tmpdir = None
    try:
        frames, meta, tmpdir = _extract_video_frames(path)
        result = {"filename": original, "size_bytes": path.stat().st_size, **meta}
        if not frames:
            result["analysis"] = "동영상에서 분석 가능한 프레임을 추출하지 못했습니다."
            return result
        try:
            result["analysis"] = vision_analyze(
                frames,
                "이 이미지들은 하나의 동영상에서 시간 순서대로 추출한 장면입니다. 전체 영상의 내용, 등장 객체, 주요 변화와 흐름을 한국어로 요약해 주세요. 오디오는 직접 들은 것이 아니므로 음성 내용은 추측하지 마세요."
            )
            result["sampled_frames"] = len(frames)
            result["vision_model_used"] = True
        except Exception as exc:
            result["analysis"] = "동영상 프레임은 추출했지만 비전 모델 분석을 실행하지 못했습니다."
            result["vision_model_used"] = False
            result["vision_error"] = str(exc)
        return result
    finally:
        path.unlink(missing_ok=True)
        if tmpdir:
            shutil.rmtree(tmpdir, ignore_errors=True)


def _read_text_file(path):
    return path.read_text(encoding="utf-8", errors="replace")


def _extract_document_text(path):
    ext = path.suffix.lower()
    if ext in {".txt", ".md", ".py", ".js", ".html", ".css", ".java", ".c", ".cpp", ".cs"}:
        return _read_text_file(path)
    if ext == ".json":
        data = json.loads(_read_text_file(path))
        return json.dumps(data, ensure_ascii=False, indent=2)
    if ext == ".csv":
        return _read_text_file(path)
    if ext == ".pdf":
        if PdfReader is None:
            raise RuntimeError("pypdf가 설치되어 있지 않습니다.")
        reader = PdfReader(str(path))
        return "\n\n".join((page.extract_text() or "") for page in reader.pages)
    if ext == ".docx":
        if Document is None:
            raise RuntimeError("python-docx가 설치되어 있지 않습니다.")
        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs)
    if ext == ".xlsx":
        if load_workbook is None:
            raise RuntimeError("openpyxl이 설치되어 있지 않습니다.")
        wb = load_workbook(str(path), read_only=True, data_only=True)
        chunks = []
        for ws in wb.worksheets[:10]:
            chunks.append(f"[시트: {ws.title}]")
            for row in ws.iter_rows(values_only=True):
                chunks.append("\t".join("" if v is None else str(v) for v in row))
                if sum(len(x) for x in chunks) > 120000:
                    break
        return "\n".join(chunks)
    raise ValueError("지원하지 않는 문서 형식입니다.")


def analyze_document_upload(file_storage):
    path, original = _safe_upload(file_storage, ALLOWED_DOC_EXT)
    try:
        text = _extract_document_text(path)
        clipped = text[:100000]
        result = {
            "filename": original,
            "extension": path.suffix.lower(),
            "size_bytes": path.stat().st_size,
            "characters_extracted": len(text),
        }
        if not clipped.strip():
            result["analysis"] = "텍스트를 추출하지 못했습니다."
            return result

        router = PickLLMRouter()
        prompt = f"""다음 파일 내용을 한국어로 분석해 주세요.
1. 핵심 요약
2. 중요한 항목
3. 주의하거나 확인할 부분
4. 사용자가 다음에 할 만한 작업

파일명: {original}

내용:
{clipped}
"""
        result["analysis"] = router.generate(prompt)
        return result
    finally:
        path.unlink(missing_ok=True)
